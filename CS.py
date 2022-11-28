import numpy as np #导入numpy库
import  pandas  as pd
import math

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

import xlwt  # 需要的模块


def MIS(fin,fout): # Interval normalization
    # df=pd.read_excel('1.xlsx',sheet_name='student')#可以通过sheet_name来指定读取的表单
    df = pd.read_excel(fin)
    # 生成excel的方法，声明excel
    #xlsx = openpyxl.Workbook()
    #sheet = xlsx.create_sheet(index=0)
    #sheet.cell(1, 1).value = ILLEGAL_CHARACTERS_RE.sub(r'', '')
    xls = xlwt.Workbook()
    sheet = xls.add_sheet('sheet1', cell_overwrite_ok=True)
    sheet.write(0, 0, "")

    K = 24 #特征数量
    row = len(df)  # 行
    sum = [0] * K
    mean = [0] * K
    error = [0] * K
    e2 = [0] * K

    #print(row)
    for i in range(0,row):
        for j in range(0,K):
            sum[j] += (df.iloc[i].values[2*j]+df.iloc[i].values[2*j+1])/2
    for i in range(0,row):
        for j in range(0, K):
            temp1 = df.iloc[i].values[2*j]
            temp2 = df.iloc[i].values[2*j+1]
            mean[j] = sum[j]/row
            error[j] += np.power((temp1+temp2)/2 - mean[j],2)
    for j in range(0, K):
        e2[j] = math.pow(1.0 / row * error[j], 0.5)
    #print(e2)
    for i in range(0,row):
        sheet.write(i+1, K, df.iloc[i].values[2*K])
        #sheet.cell(i + 2, 1).value = df.iloc[i].values[0]
        for j in range(0, K):
            if e2[j] == 0:
                sheet.write(i + 1, j, 0)
            else:
                object = ((df.iloc[i].values[2*j]+df.iloc[i].values[2*j+1])/2 - mean[j]) / e2[j]
                sheet.write(i+1, j, object)  # x单元格纬度，i 单元格经度

    xls.save(fout)  # 保存xls文件  # 保存xls文件
    #xlsx.save(fout)
if __name__ == "__main__":
    fin = "data/wall/data-Inv.xls"
    fout = "data/wall/data-MIS.xls"
    MIS(fin, fout)
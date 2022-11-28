import numpy as np #导入numpy库
import  pandas  as pd
import math

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

import xlwt  # 需要的模块


def MIN(fin,fout): # Interval normalization
    # df=pd.read_excel('1.xlsx',sheet_name='student')#可以通过sheet_name来指定读取的表单
    df = pd.read_excel(fin)
    # 生成excel的方法，声明excel
    #xlsx = openpyxl.Workbook()
    #sheet = xlsx.create_sheet(index=0)
    #sheet.cell(1, 1).value = ILLEGAL_CHARACTERS_RE.sub(r'', '')
    xls = xlwt.Workbook()
    sheet = xls.add_sheet('sheet1', cell_overwrite_ok=True)
    sheet.write(0, 0, "")

    K = 24 #特征数量
    row = len(df)  # 行
    mean = np.zeros((row,K))
    sum = [0]*K
    min = [0] * K
    max = [0] * K
    #print(row)
    for i in range(0,row):
        for j in range(0,K):
            if float(min[j]) >= float(df.iloc[i].values[2*j]):
                min[j] = df.iloc[i].values[2*j]
            if float(max[j]) <= float(df.iloc[i].values[2*j+1]):
                max[j] = df.iloc[i].values[2*j+1]
            mean[i][j] = (df.iloc[i].values[2*j]+df.iloc[i].values[2*j+1])/2

    for i in range(0,row):
        sheet.write(i+1, K, df.iloc[i].values[2*K])
        #sheet.cell(i + 2, 1).value = df.iloc[i].values[0]
        for j in range(0, K):
            sheet.write(i+1, j, mean[i][j]/(max[j]-min[j]))  # x单元格纬度，i 单元格经度

    xls.save(fout)  # 保存xls文件  # 保存xls文件
    #xlsx.save(fout)
if __name__ == "__main__":
    fin = "data/wall/data-Inv.xls"
    fout = "data/wall/data-MIN.xls"
    MIN(fin, fout)
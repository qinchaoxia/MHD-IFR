import numpy as np #导入numpy库
import  pandas  as pd
import math

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

import xlwt  # 需要的模块


def IN(fin,fout): # Interval normalization
    # df=pd.read_excel('1.xlsx',sheet_name='student')#可以通过sheet_name来指定读取的表单
    df = pd.read_excel(fin)
    # 生成excel的方法，声明excel
    '''
    xlsx = openpyxl.Workbook()
    sheet = xlsx.create_sheet(index=0)
    sheet.cell(1, 1).value = ILLEGAL_CHARACTERS_RE.sub(r'', '')
    '''
    xls = xlwt.Workbook()
    sheet = xls.add_sheet('sheet1', cell_overwrite_ok=True)
    sheet.write(0, 0, "")

    K = 24 #特征数量
    sum1 = [0]*K
    sum2 = [0]*K
    row = len(df)  # 行
    #print(row)
    for i in range(0,row):
        for j in range(0,K):
            sum1[j] += df.iloc[i].values[2*j]
            sum2[j] += df.iloc[i].values[2*j+1]

    for i in range(0,row):
        sheet.write(i+1, 2*K, df.iloc[i].values[2*K])
        #sheet.cell(i + 2, 2*K+1).value = df.iloc[i].values[2*K]
        for j in range(0, K):
            object1 = df.iloc[i].values[2*j] / sum1[j]
            object2 = df.iloc[i].values[2*j] / sum2[j]
            object3 = df.iloc[i].values[2*j+1] / sum1[j]
            object4 = df.iloc[i].values[2*j+1] / sum2[j]
            sheet.write(i+1, 2*j, min(object1,object2,object3,object4))  # x单元格纬度，i 单元格经度
            sheet.write(i+1, 2*j+1, max(object1,object2,object3,object4))
            #sheet.cell(i + 2, 2*j+1).value = min(object1,object2,object3,object4)
            #sheet.cell(i + 2, 2*j+2).value = max(object1, object2, object3, object4)

    xls.save(fout)  # 保存xls文件  # 保存xls文件
    #xlsx.save(fout)
if __name__ == "__main__":
    fin = "data/wall/data-Inv.xls"
    fout = "data/wall/data-IN.xls"
    IN(fin, fout)
# -*- encoding: utf-8 -*-
import random
import time

import xlwt  # 需要的模块

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

import math
import numpy as np
import  pandas  as pd
def zscore(fin,fout,col):
    """
    :文本转换成xls的函数
    :param filename txt文本文件名称、
    :param xlsname 表示转换后的excel文件名
    """
    try:
        df = pd.read_excel(fin)
        '''
        xls = xlwt.Workbook()
        sheet = xls.add_sheet('sheet1', cell_overwrite_ok=True)
        sheet.write(0, 0, '')
        '''
        xlsx = openpyxl.Workbook()
        sheet = xlsx.create_sheet(index=0)
        sheet.cell(1, 1).value = ILLEGAL_CHARACTERS_RE.sub(r'', '')

        M = 20
        t =8
        K =col
        row = len(df)

        for m in range(0,M): # 按信息源循环，读取文本文件
            num = np.zeros((t, K))  # 每个类别的对象个数
            sum = np.zeros((t, K))  # 每个类别的每个属性之和
            mean = np.zeros((t, K))  # 每个类别的每个属性均值
            E2 = np.zeros((t, K))  # 每个类别的每个属性方差
            for i in range(0,row,M):
                label = df.iloc[i+m].values[0]
                for j in range(1,K):
                    item = df.iloc[i+m].values[j]
                    if label==1:
                        num[0][j]=num[0][j]+1
                        sum[0][j] = sum[0][j] + item
                    if label==2:
                        num[1][j]=num[1][j]+1
                        sum[1][j] = sum[1][j] + item
                    if label==3:
                        num[2][j]=num[2][j]+1
                        sum[2][j] = sum[2][j] + item
                    if label==4:
                        num[3][j]=num[3][j]+1
                        sum[3][j] = sum[3][j] + item
                    if label==5:
                        num[4][j]=num[4][j]+1
                        sum[4][j] = sum[4][j] + item
                    if label==6:
                        num[5][j]=num[5][j]+1
                        sum[5][j] = sum[5][j] + item
                    if label==7:
                        num[6][j]=num[6][j]+1
                        sum[6][j] = sum[6][j] + item
                    if label==8:
                        num[7][j]=num[7][j]+1
                        sum[7][j] = sum[7][j] + item
            for i in range(0,row,M):
                label = df.iloc[i + m].values[0]
                for j in range(1, K):
                    item = df.iloc[i+m].values[j]
                    if label == 1:
                        mean[0][j] = sum[0][j] / num[0][j]
                        e2 = math.pow(item - mean[0][j], 2)
                        E2[0][j] = E2[0][j] + e2
                    if label == 2:
                        mean[1][j] = sum[1][j] / num[1][j]
                        e2 = math.pow(item - mean[1][j], 2)
                        E2[1][j] = E2[1][j] + e2
                    if label == 3:
                        mean[2][j] = sum[2][j] / num[2][j]
                        e2 = math.pow(item - mean[2][j], 2)
                        E2[2][j] = E2[2][j] + e2
                    if label == 4:
                        mean[3][j] = sum[3][j] / num[3][j]
                        e2 = math.pow(item - mean[3][j], 2)
                        E2[3][j] = E2[3][j] + e2
                    if label == 5:
                        mean[4][j] = sum[4][j] / num[4][j]
                        e2 = math.pow(item - mean[4][j], 2)
                        E2[4][j] = E2[4][j] + e2
                    if label == 6:
                        mean[5][j] = sum[5][j] / num[5][j]
                        e2 = math.pow(item - mean[5][j], 2)
                        E2[5][j] = E2[5][j] + e2
                    if label == 7:
                        mean[6][j] = sum[6][j] / num[6][j]
                        e2 = math.pow(item - mean[6][j], 2)
                        E2[6][j] = E2[6][j] + e2
                    if label == 8:
                        mean[7][j] = sum[7][j] / num[7][j]
                        e2 = math.pow(item - mean[7][j], 2)
                        E2[7][j] = E2[7][j] + e2
            for i in range(0, row, M):
                #print(i+m+1)
                #sheet.write(i+m+1, K-1, df.iloc[i+m].values[0])  # 0是类别
                sheet.cell(i+m+2, K).value = df.iloc[i+m].values[0]
                label = df.iloc[i + m].values[0]
                for j in range(1, K):
                    item = df.iloc[i + m].values[j]
                    if label == 1:
                        e = math.pow(1.0 / num[0][j] * E2[0][j], 0.5)
                        if e == 0:
                            item = 0
                        else:
                            item = (item - mean[0][j])/e
                    if label == 2:
                        e = math.pow(1.0 / num[1][j] * E2[1][j], 0.5)
                        if e == 0:
                            item = 0
                        else:
                            item = (item - mean[1][j])/e
                    if label == 3:
                        e = math.pow(1.0 / num[2][j] * E2[2][j], 0.5)
                        if e == 0:
                            item = 0
                        else:
                            item = (item - mean[2][j])/e
                    if label == 4:
                        e = math.pow(1.0 / num[3][j] * E2[3][j], 0.5)
                        if e == 0:
                            item = 0
                        else:
                            item = (item - mean[3][j])/e
                    if label == 5:
                        e = math.pow(1.0 / num[4][j] * E2[4][j], 0.5)
                        if e == 0:
                            item = 0
                        else:
                            item = (item - mean[4][j])/e
                    if label == 6:
                        e = math.pow(1.0 / num[5][j] * E2[5][j], 0.5)
                        if e == 0:
                            item = 0
                        else:
                            item = (item - mean[5][j])/e
                    if label == 7:
                        e = math.pow(1.0 / num[6][j] * E2[6][j], 0.5)
                        if e == 0:
                            item = 0
                        else:
                            item = (item - mean[6][j])/e
                    if label == 8:
                        e = math.pow(1.0 / num[7][j] * E2[7][j], 0.5)
                        if e == 0:
                            item = 0
                        else:
                            item = (item - mean[7][j])/e
                    #sheet.write(i+m+1, j-1, item)
                    sheet.cell(i+m+2, j).value = item


        xlsx.save(fout)  # 保存xls文件  # 保存xls文件
    except:
        raise
if __name__ == "__main__":

    fin = "../data/wall/data.xls"
    fout = "../data/wall/data-S.xlsx"
    col = 1 + 24  # type+attributes
    start = time.time()
    zscore(fin,fout,col)
    end = time.time()
    print(end - start)

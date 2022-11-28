# -*- encoding: utf-8 -*-
import random

import xlwt  # 需要的模块

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import math
import numpy as np
import  pandas  as pd
def multiple(fin,fout):
    """
    :文本转换成xls的函数
    :param filename txt文本文件名称、
    :param xlsname 表示转换后的excel文件名
    """
    try:
        f = open(fin)
        '''
        xls = xlwt.Workbook()
        sheet = xls.add_sheet('sheet1', cell_overwrite_ok=True)
        sheet.write(0, 0, '')
        '''
        xlsx = openpyxl.Workbook()
        sheet = xlsx.create_sheet(index=0)
        sheet.cell(1, 1).value = ILLEGAL_CHARACTERS_RE.sub(r'', '')

        x = 1  # 换行
        M = 20
        t = 8  # type
        K = 1+24 # type + attributes

        num = np.zeros((t, K))  # 每个类别的对象个数
        sum = np.zeros((t, K))  # 每个类别的每个属性之和
        mean = np.zeros((t, K))  # 每个类别的每个属性均值
        E2 = np.zeros((t, K))  # 每个类别的每个属性方差
        while True:
            # 按行循环，读取文本文件
            line = f.readline()
            if not line:
                break  # 如果没有内容，则退出循环
            for i in range(len(line.split(','))):
                if i>0: # 1-12是属性
                    item = float(line.split(',')[i])
                    if (line.split(',')[0])=='1':
                        num[0][i]=num[0][i]+1
                        sum[0][i] = sum[0][i] + item
                    if (line.split(',')[0])=='2':
                        num[1][i]=num[1][i]+1
                        sum[1][i] = sum[1][i] + item
                    if (line.split(',')[0])=='3':
                        num[2][i]=num[2][i]+1
                        sum[2][i] = sum[2][i] + item
                    if (line.split(',')[0])=='4':
                        num[3][i]=num[3][i]+1
                        sum[3][i] = sum[3][i] + item
                    if (line.split(',')[0])=='5':
                        num[4][i]=num[4][i]+1
                        sum[4][i] = sum[4][i] + item
                    if (line.split(',')[0])=='6':
                        num[5][i]=num[5][i]+1
                        sum[5][i] = sum[5][i] + item
                    if (line.split(',')[0])=='7':
                        num[6][i]=num[6][i]+1
                        sum[6][i] = sum[6][i] + item
                    if (line.split(',')[0])=='8':
                        num[7][i]=num[7][i]+1
                        sum[7][i] = sum[7][i] + item
        f.close()
        f = open(fin)
        while True:
            # 按行循环，读取文本文件
            line = f.readline()
            if not line:
                break  # 如果没有内容，则退出循环
            for i in range(len(line.split(','))):
                if i>0:  # 1-12是属性
                    item = float(line.split(',')[i])
                    if (line.split(',')[0])=='1':
                        mean[0][i]= sum[0][i] / num[0][i]
                        e2= math.pow(item-mean[0][i],2)
                        E2[0][i]=E2[0][i]+e2
                    if (line.split(',')[0])=='2':
                        mean[1][i] = sum[1][i] / num[1][i]
                        e2 = math.pow(item - mean[1][i], 2)
                        E2[1][i] = E2[1][i] + e2
                    if (line.split(',')[0])=='3':
                        mean[2][i] = sum[2][i] / num[2][i]
                        e2 = math.pow(item - mean[2][i], 2)
                        E2[2][i] = E2[2][i] + e2
                    if (line.split(',')[0])=='4':
                        mean[3][i]= sum[3][i] / num[3][i]
                        e2 = math.pow(item-mean[3][i],2)
                        E2[3][i]=E2[3][i]+e2
                    if (line.split(',')[0])=='5':
                        mean[4][i]= sum[4][i] / num[4][i]
                        e2 = math.pow(item-mean[4][i],2)
                        E2[4][i]=E2[4][i]+e2
                    if (line.split(',')[0])=='6':
                        mean[5][i] = sum[5][i] / num[5][i]
                        e2 = math.pow(item - mean[5][i], 2)
                        E2[5][i] = E2[5][i] + e2
                    if (line.split(',')[0])=='7':
                        mean[6][i] = sum[6][i] / num[6][i]
                        e2 = math.pow(item - mean[6][i], 2)
                        E2[6][i] = E2[6][i] + e2
                    if (line.split(',')[0])=='8':
                        mean[7][i]= sum[7][i] / num[7][i]
                        e2 = math.pow(item-mean[7][i],2)
                        E2[7][i]=E2[7][i]+e2

        f.close()
        f = open(fin)
        while True:
            # 按行循环，读取文本文件
            line = f.readline()
            if not line:
                break  # 如果没有内容，则退出循环
            for m in range(M):
                for i in range(len(line.split(','))):
                    r = random.uniform(-1, 1)
                    #print(r)
                    if i == 0:  # 0是类别
                        item = (line.split(',')[i])
                        #sheet.write(x, i, item)  # x单元格纬度，i 单元格经度
                        sheet.cell(x+1, i+1).value = item
                    if i > 0:  # 1-12是属性
                        item = float(line.split(',')[i])
                        if (line.split(',')[0]) == '1':
                            e = math.pow(1.0 / num[0][i] * E2[0][i], 0.5)
                            item = item + r * e
                        if (line.split(',')[0]) == '2':
                            e = math.pow(1.0 / num[1][i] * E2[1][i], 0.5)
                            item = item - r * e
                        if (line.split(',')[0]) == '3':
                            e = math.pow(1.0 / num[2][i] * E2[2][i], 0.5)
                            item = item + r * e
                        if (line.split(',')[0]) == '4':
                            e = math.pow(1.0 / num[3][i] * E2[3][i], 0.5)
                            item = item + r * e
                        if (line.split(',')[0]) == '5':
                            e = math.pow(1.0 / num[4][i] * E2[4][i], 0.5)
                            item = item + r * e
                        if (line.split(',')[0]) == '6':
                            e = math.pow(1.0 / num[5][i] * E2[5][i], 0.5)
                            item = item + r * e
                        if (line.split(',')[0]) == '7':
                            e = math.pow(1.0 / num[6][i] * E2[6][i], 0.5)
                            item = item + r * e
                        if (line.split(',')[0]) == '8':
                            e = math.pow(1.0/num[7][i]*E2[7][i],0.5)
                            item = item + r * e
                        #sheet.write(x, i, item)
                        sheet.cell(x+1, i+1).value = item
                x += 1  # excel另起一行
        f.close()
        xlsx.save(fout)  # 保存xls文件  # 保存xls文件
    except:
        raise

if __name__ == "__main__":

    fin = "../data/wall/wall.txt"
    fout = "../data/wall/data.xls"
    multiple(fin,fout)

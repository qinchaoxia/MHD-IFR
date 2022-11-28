# -*- encoding: utf-8 -*-
import xlwt  # 需要的模块
import xlrd

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

import pandas as pd
import glob, os
import csv
import codecs
def txt_xls(txtname, xlsname):
    try:
        f = open(txtname)
        xls = xlwt.Workbook()
        sheet = xls.add_sheet('sheet1', cell_overwrite_ok=True)
        x = 0
        while True:
            # 按行循环，读取文本文件
            line = f.readline()
            if not line:
                break  # 如果没有内容，则退出循环
            for i in range(len(line.split(','))):
                item = line.split(',')[i]
                sheet.write(x, i, item)  # x单元格纬度，i 单元格经度
            x += 1  # excel另起一行
        f.close()
        xls.save(xlsname)  # 保存xls文件
    except:
        raise
def txt_xlsx(txtname, xlsname):
    try:
        f = open(txtname)
        xlsx = openpyxl.Workbook()
        sheet = xlsx.create_sheet(index=0)
        x = 1
        while True:
            # 按行循环，读取文本文件
            line = f.readline()
            if not line:
                break  # 如果没有内容，则退出循环
            for i in range(len(line.split(','))):
                item = line.split(',')[i]
                #sheet.write(x, i, item)  # x单元格纬度，i 单元格经度
                sheet.cell(x, i+1).value = item
            x += 1  # excel另起一行
        f.close()
        xlsx.save(xlsname)  # 保存xls文件
    except:
        raise
def strs(row):
        try:
            values = "";
            for i in range(len(row)):
                if i == len(row) - 1:
                    #values = values + str(row[i])+'\n'
                    values = values + str(row[i])
                else:
                    # 使用“，”逗号作为分隔符
                    values = values + str(row[i]) + ","
            return values
        except:
            raise
def xls_txt(xlsname, txtname):
        try:
            data = xlrd.open_workbook(xlsname)
            sqlfile = open(txtname, "w")
            table = data.sheets()[0]  # 表头
            nrows = table.nrows  # 行数
            # 如果不需跳过表头，则将下一行中1改为0
            for ronum in range(0, nrows):
                row = table.row_values(ronum)
                values = strs(row)  # 条用函数，将行数据拼接成字符串
                sqlfile.writelines(values)  # 将字符串写入新文件
            sqlfile.close()  # 关闭写入的文件
        except:
            pass

if __name__ == "__main__":

    fin = "../data/test/data.txt"
    fout = "../data/test/data.xls"
    txt_xls(fin, fout)
    #txt_xlsx(fin, fout)
    #xls_txt(fout, fin)
    '''
    path = r'../data/AReM'
    path2 = r'../data/AReM/xls'
    ff = glob.glob(os.path.join(path, "*.xls"))
    x = 0
    for fin in ff:
        fout = path2 + str(x) + ".xls"
        x += 1
        txt_xls(fin, fout)
    '''


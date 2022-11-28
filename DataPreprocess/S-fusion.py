# -*- encoding: utf-8 -*-
import random
import time

import xlwt  # 需要的模块

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

import math
import numpy as np
import  pandas  as pd
def fuzzy(fin,fout,col):
    """
    :文本转换成xls的函数
    :param filename txt文本文件名称、
    :param xlsname 表示转换后的excel文件名
    """
    try:
        df = pd.read_excel(fin)

        xls = xlwt.Workbook()
        sheet = xls.add_sheet('sheet1', cell_overwrite_ok=True)
        sheet.write(0, 0, '')
        '''
        xlsx = openpyxl.Workbook()
        sheet = xlsx.create_sheet(index=0)
        sheet.cell(1, 1).value = ILLEGAL_CHARACTERS_RE.sub(r'', '')
        '''
        M = 20
        K = col # attributes
        row = len(df)

        for i in range(0,row,M): # 按行循环，读取文本文件
            sum = [0] * K
            mean = [0] * K
            e2 = [0] * K
            sheet.write(i // M + 1, K*2, df.iloc[i].values[K])  # 0是类别
            #sheet.cell(i//M+2, K*2+1).value = df.iloc[i].values[K]
            for m in range(M):
                for j in range(K):
                    sum[j] += df.iloc[i+m].values[j]
            mean[j] = sum[j]/M
            for m in range(M):
                for j in range(K):
                    e2[j] += (df.iloc[i+m].values[j]-mean[j])**2
            for j in range(K):
                e = np.power(e2[j]/M,0.5)
                z = 1
                sheet.write(i // M + 1, 2 * j, mean[j]-e)
                sheet.write(i // M + 1, 2 * j+1, mean[j]+e)
                #sheet.cell(i//M+2, j*2+1).value = mean[j]-z*e
                #sheet.cell(i//M +2, j*2+2).value = mean[j]+z*e


        xls.save(fout)  # 保存xls文件  # 保存xls文件
    except:
        raise
if __name__ == "__main__":
    fin = "../data/trip/data-S.xls"
    fout = "../data/trip/data-S-Inv.xls"
    col = 10 # attributes
    start = time.time()
    fuzzy(fin,fout,col)
    end = time.time()
    print(end-start)
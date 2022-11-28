import time

import numpy as np #导入numpy库
import  pandas  as pd
import math

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

import xlwt  # 需要的模块


def IS(hist_mean,hist_var,fin,fout,col,hist_mean2,hist_var2): # Interval normalization
    # df=pd.read_excel('1.xlsx',sheet_name='student')#可以通过sheet_name来指定读取的表单
    df = pd.read_excel(fin)
    hist_mean = pd.read_excel(hist_mean)
    hist_var = pd.read_excel(hist_var)
    # 生成excel的方法，声明excel
    '''
    xlsx = openpyxl.Workbook()
    sheet = xlsx.create_sheet(index=0)
    sheet.cell(1, 1).value = ILLEGAL_CHARACTERS_RE.sub(r'', '')
    '''
    xls1 = xlwt.Workbook()
    sheet1 = xls1.add_sheet('sheet1', cell_overwrite_ok=True)
    sheet1.write(0, 0, "")

    xls2 = xlwt.Workbook()
    sheet2 = xls2.add_sheet('sheet1', cell_overwrite_ok=True)
    sheet2.write(0, 0, "")

    xls3 = xlwt.Workbook()
    sheet3 = xls3.add_sheet('sheet1', cell_overwrite_ok=True)
    sheet3.write(0, 0, "")

    K = col
    sum1 = [0]*K
    sum2 = [0]*K
    mean1 = [0] * K
    mean2 = [0] * K
    error1 = [0]*K
    error2 = [0]*K
    e1 = [0] * K
    e2 = [0] * K
    row = len(df)  # 行
    #print(row)
    for i in range(0,row):
        for j in range(0,K):
            sum1[j] += df.iloc[i].values[2*j]
            sum2[j] += df.iloc[i].values[2*j+1]
    for i in range(0,row):
        for j in range(0, K):
            temp1 = df.iloc[i].values[2*j]
            temp2 = df.iloc[i].values[2*j+1]
            mean1[j] = sum1[j]/row
            mean2[j] = sum2[j]/row
            if (temp1 - mean2[j]) * (temp2 - mean1[j]) >= 0:
                a1 = np.power(temp1 - mean2[j], 2)
                a2 = np.power(temp2 - mean1[j], 2)
                error1[j] += min(a1, a2)
                error2[j] += max(a1, a2)
            if (temp1 - mean2[j]) * (temp2 - mean1[j]) < 0:
                a1 = np.power(temp1 - mean2[j], 2)
                a2 = np.power(temp2 - mean1[j], 2)
                error1[j] += 0
                error2[j] += max(a1, a2)
    for j in range(0, K):
        mean1[j] = (mean1[j]*row+hist_mean.iloc[0].values[2 * j]*row)/(row+row)
        mean2[j] = (mean2[j] * row + hist_mean.iloc[0].values[2 * j + 1] * row) / (row + row)
        sheet2.write(1, 2 * j, mean1[j])
        sheet2.write(1, 2 * j + 1, mean2[j])
        error1[j] = (error1[j] * row + hist_var.iloc[0].values[2 * j] * row) / (row + row)
        error2[j] = (error2[j] * row + hist_var.iloc[0].values[2 * j + 1] * row) / (row + row)
        sheet3.write(1, 2 * j, error1[j])  # x单元格纬度，i 单元格经度
        sheet3.write(1, 2 * j + 1, error2[j])
        e1[j] = math.pow(1.0 / row * error1[j], 0.5)
        e2[j] = math.pow(1.0 / row * error2[j], 0.5)
    for i in range(0,row):
        sheet1.write(i+1, 2*K, df.iloc[i].values[2*K])
        #sheet1.cell(i + 2, 2*K+1).value = df.iloc[i].values[2*K]
        for j in range(0, K):
            if e1[j] == 0:
                object2 = (df.iloc[i].values[2 * j] - mean2[j]) / e2[j]
                object4 = (df.iloc[i].values[2 * j + 1] - mean1[j]) / e2[j]
                sheet1.write(i + 1, 2 * j, min(0,object2, object4))  # x单元格纬度，i 单元格经度
                sheet1.write(i + 1, 2 * j + 1, max(0,object2, object4))
                #sheet1.cell(i + 2, 2 * j + 1).value = 0
                #sheet1.cell(i + 2, 2 * j + 2).value = max(object2, object4)
            elif e2[j] == 0:
                sheet1.write(i + 1, 2 * j, 0)  # x单元格纬度，i 单元格经度
                sheet1.write(i + 1, 2 * j + 1, 0)
                #sheet1.cell(i + 2, 2 * j + 1).value = 0
                #sheet1.cell(i + 2, 2 * j + 2).value = 0
            else:
                object1 = (df.iloc[i].values[2 * j] - mean2[j]) / e1[j]
                object2 = (df.iloc[i].values[2 * j] - mean2[j]) / e2[j]
                object3 = (df.iloc[i].values[2 * j + 1] - mean1[j]) / e1[j]
                object4 = (df.iloc[i].values[2 * j + 1] - mean1[j]) / e2[j]
                sheet1.write(i+1, 2*j, min(object1,object2,object3,object4))  # x单元格纬度，i 单元格经度
                sheet1.write(i+1, 2*j+1, max(object1,object2,object3,object4))
                #sheet1.cell(i + 2, 2*j+1).value = min(object1,object2,object3,object4)
                #sheet1.cell(i + 2, 2*j+2).value = max(object1, object2, object3, object4)

    xls1.save(fout)  # 保存xls文件  # 保存xls文件
    xls2.save(hist_mean2)
    xls3.save(hist_var2)
if __name__ == "__main__":
    fin = "data/iris/data-Inv.xls"
    hist_mean = "data/iris/data-IS-mean.xls"
    hist_var = "data/iris/data-IS-var.xls"

    fout = "data/iris/data-IS2.xls"
    hist_mean2 = "data/iris/data-IS-mean2.xls"
    hist_var2 = "data/iris/data-IS-var2.xls"
    col = 4  # 特征数量
    start = time.time()
    IS(hist_mean,hist_var,fin, fout,col,hist_mean2,hist_var2)
    end = time.time()
    print(end-start)
